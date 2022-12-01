import openpyxl as op

def checkExcel():
    pdf1 = op.load_workbook("C:/Users/GHOSTLaycoo/Desktop/营销场景二/data/temp1.xlsx")
    pdf2 = op.load_workbook("C:/Users/GHOSTLaycoo/Desktop/营销场景二/data/temp3.xlsx")

    table_1 = []
    table_2 = []

    ws_1 = pdf1.active
    ws_2 = pdf2.active
    valuesByRow_1 = tuple(ws_1.values)
    valuesByRow_2 = tuple(ws_2.values)
    cellValueList_1 = read_value_by_column_header(table_1, valuesByRow_1)
    cellValueList_2 = read_value_by_column_header(table_2, valuesByRow_2)
    is_in_2 = True
    for row_2 in valuesByRow_2[1:]:
        x = list(row_2)
        for row_1 in valuesByRow_1[1:]:
            y = list(row_1)
            if(row_2[cellValueList_2.get('银行单据号码')] == row_1[cellValueList_1.get('流水号')]):
                if(str(row_2[cellValueList_2.get('入账金额')])!=str(row_1[cellValueList_1.get('收入')]).replace(',','')):
                    x.append("存在数据，但金额有差异")
                    y.append("存在数据，但金额有差异")
                    table_2.append(x)
                    table_1.append(y)
                if(is_in_2):
                    is_in_2 = False
        if(is_in_2):
            x.append("另外系统或excel,不存在该数据")
            table_2.append(x)
        else:
            is_in_2 = True

    is_in_1 = True
    for row_1 in valuesByRow_1[1:]:
        y = list(row_1)
        for row_2 in valuesByRow_2[1:]:
            if(row_2[cellValueList_2.get('银行单据号码')] == row_1[cellValueList_1.get('流水号')]):
                if(is_in_1):
                    is_in_1 = False
        if(is_in_1):
            y.append("另外系统或excel,不存在该数据")
            table_1.append(y)
        else:
            is_in_1 = True

    createExcel(table_1, table_2)


def createExcel(table_1, table_2):
    wk = op.Workbook()
    excelSheet = wk.create_sheet("excelData", index=0)
    systemSheet = wk.create_sheet("systemData", index=1)
    importData(excelSheet, table_1)
    importData(systemSheet, table_2)
    wk.save("C:/Users/GHOSTLaycoo/Desktop/营销场景二/data/result.xlsx")

def importData(sheet, table):
    for row in table:
        sheet.append(row)



# 以表头生成字典模型，方便后面获取数据
def read_value_by_column_header(table, valuesByRow):
    firstRow = valuesByRow[0]
    title = list(firstRow)
    title.append("备注")
    table.append(title)
    cellValueList = {}
    for columnHeader in firstRow:
        for i, v in enumerate(firstRow):
            if v == columnHeader:
                cellValueList[columnHeader] = i
                break
    return cellValueList



if __name__ == '__main__':
    checkExcel()